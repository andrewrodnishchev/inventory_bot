import os
import sqlite3
import logging
from datetime import datetime
from aiogram import Bot, Dispatcher, types, F, Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.webhook.aiohttp_server import SimpleRequestHandler, setup_application
from aiohttp import web
from dotenv import load_dotenv
from openpyxl import Workbook

load_dotenv("token.env")

bot_token = os.getenv("BOT_TOKEN")
bot = Bot(token=bot_token)
dp = Dispatcher()
router = Router()
dp.include_router(router)

# Категории для разных бланков
BLANKS = {
    "bar": [
        "БАР - ПРОЧЕЕ товар",
        "БЕЗАЛКОГОЛЬНЫЕ НАПИТКИ**",
        "ЧАЙ/КОФЕ товар",
        "КОНДИТЕРКА/ВЫПЕЧКА",
        "КОНСЕРВАЦИЯ",
        "СОУСЫ",
        "СПЕЦИИ",
        "СЫПУЧИЕ",
        "ЧАЙ, КОФЕ",
        "ЗЕЛЕНЬ",
        "МОЛОЧНЫЕ ПРОДУКТЫ",
        "ОВОЩИ СВЕЖИЕ",
        "ФРУКТЫ",
        "ЯГОДЫ С/М",
        "НАПИТКИ БЛ ПФ"
    ],
    "alcohol": [
        "ВИНО БЕЛОЕ",
        "ВИНО КРАСНОЕ",
        "ВИНО ОРАНЖЕВОЕ",
        "ВЕРМУТ",
        "ШАМПАНСКОЕ/ИГРИСТОЕ",
        "ВИСКИ",
        "ВОДКА",
        "ГОРЬКИЕ НАСТОЙКИ",
        "ДЖИН",
        "КАЛЬВАДОС",
        "КОНЬЯК/АРМАНЬЯК",
        "ЛИКЕРЫ",
        "ПИВО РАЗЛИВНОЕ",
        "ПОРТВЕЙН",
        "ПОРТО/ХЕРЕС",
        "РОМ",
        "ТЕКИЛА",
        "ПИВО БУТЫЛКА",
        "БЕЗАЛКОГОЛЬНОЕ ПИВО"
    ]
}

# База данных
conn = sqlite3.connect("inventory.db", check_same_thread=False)
cursor = conn.cursor()
cursor.execute("DROP TABLE IF EXISTS inventory")
cursor.execute('''
CREATE TABLE IF NOT EXISTS inventory (
    user_id INTEGER,
    username TEXT,
    category TEXT,
    name TEXT,
    quantity REAL,
    UNIQUE(user_id, category, name)
)
''')
conn.commit()

class InventoryState(StatesGroup):
    waiting_for_item_data = State()
    waiting_for_edit_input = State()  # ожидание нового ввода для редактирования

@dp.message(F.text == "/start")
async def start(message: types.Message, state: FSMContext):
    # Возврат в главное меню
    await state.clear()
    builder = InlineKeyboardBuilder()
    # Основные функции инвентаризации
    builder.button(text="🍸 Бар", callback_data="blank_bar")
    builder.button(text="🥃 Алкоголь", callback_data="blank_alcohol")
    builder.button(text="📊 Excel (личный)", callback_data="generate_excel")
    builder.button(text="📊 Excel (общий)", callback_data="generate_excel_all")
    # Редактирование и удаление записей
    builder.button(text="✏️ Редактировать запись", callback_data="edit_menu")
    builder.button(text="❌ Удалить запись", callback_data="delete_menu")
    builder.button(text="🧹 Очистить", callback_data="clear_data")
    builder.button(text="📖 Инструкция", callback_data="instruction")
    builder.adjust(2, 2)
    await message.answer("Выберите тип инвентаризации:", reply_markup=builder.as_markup())

@dp.callback_query(F.data == "instruction")
async def show_instruction(callback: types.CallbackQuery):
    instruction_text = (
        "📚 *Инструкция по работе с ботом*\n\n"
        "1. Выберите тип инвентаризации (Бар или Алкоголь)\n"
        "2. Выберите нужную категорию из списка\n"
        "3. Вводите данные в формате:\n"
        "   `<Название товара> <Количество>`\n"
        "   Например: _Виски Джек Дэниелс 3.5_\n"
        "4. В одном сообщении МОЖНО вводить несколько товаров\n"
        "5. Используйте кнопку 📊 для генерации Excel-отчета\n"
        "6. Кнопка 🧹 очистит все ваши данные\n\n"
        "✏️ Для редактирования записи выберите «Редактировать запись»\n"
        "❌ Для удаления записи выберите «Удалить запись» и подтвердите удаление."
    )
    builder = InlineKeyboardBuilder()
    builder.button(text="◀️ Назад", callback_data="back_to_main")
    await callback.message.answer(instruction_text, parse_mode="Markdown", reply_markup=builder.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("blank_"))
async def select_blank(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    blank_type = callback.data.split("_")[1]
    await state.update_data(current_blank=blank_type)
    await show_categories(callback, blank_type)

async def show_categories(callback: types.CallbackQuery, blank_type: str):
    categories = BLANKS[blank_type]
    builder = InlineKeyboardBuilder()
    for category in categories:
        builder.button(text=category, callback_data=f"category_{category}")
    builder.button(text="🔙 Назад", callback_data="back_to_main")
    builder.adjust(1)
    new_text = f"Выберите категорию ({'Бар' if blank_type == 'bar' else 'Алкоголь'}):"
    new_markup = builder.as_markup()
    try:
        if callback.message.text != new_text or callback.message.reply_markup.to_json() != new_markup.to_json():
            await callback.message.edit_text(new_text, reply_markup=new_markup)
        else:
            await callback.answer()
    except Exception as e:
        logging.error(f"Ошибка обновления меню: {e}")
        await callback.message.answer(new_text, reply_markup=new_markup)

@dp.callback_query(F.data == "back_to_main")
async def back_to_main(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    await start(callback.message, state)

@dp.callback_query(F.data == "clear_data")
async def clear_data(callback: types.CallbackQuery):
    cursor.execute("DELETE FROM inventory WHERE user_id = ?", (callback.from_user.id,))
    conn.commit()
    await callback.answer("Данные очищены!✅", show_alert=True)

@dp.callback_query(F.data.startswith("category_"))
async def select_category(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    category = callback.data.split("_", 1)[1]
    await state.update_data(category=category)
    await state.set_state(InventoryState.waiting_for_item_data)
    await callback.message.edit_text(
        f"Категория: {category}\nВведите данные в формате:\n<Название> <Количество>\nПример: Ром 5\n"
        "Можно использовать дробные числа через . или , (например: 2.5 или 3,0)"
    )

@dp.message(InventoryState.waiting_for_item_data)
async def process_item_data(message: types.Message, state: FSMContext):
    try:
        data = await state.get_data()
        category = data["category"]
        username = message.from_user.username or message.from_user.full_name
        valid_count = 0
        items = message.text.strip().split('\n')
        for item in items:
            parts = item.rsplit(" ", 1)
            if len(parts) != 2:
                await message.answer(f"❌ Неверный формат: '{item}'")
                continue
            name, quantity_str = parts[0].strip(), parts[1].replace(',', '.')
            try:
                quantity = float(quantity_str)
            except ValueError:
                await message.answer(f"❌ Неверное количество: '{item}'")
                continue
            try:
                cursor.execute('''
                    INSERT INTO inventory (user_id, username, category, name, quantity)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(user_id, category, name) DO UPDATE SET quantity = quantity + excluded.quantity
                ''', (message.from_user.id, username, category, name, quantity))
                valid_count += 1
            except sqlite3.Error as e:
                logging.error(f"Ошибка БД: {e}")
                await message.answer(f"❌ Ошибка сохранения: {item}")
        conn.commit()
        if valid_count > 0:
            await message.answer(f"✅ Сохранено: {valid_count} позиций")
        else:
            await message.answer("❌ Нет валидных данных")
    except Exception as e:
        logging.error(f"Ошибка: {e}")
        await message.answer(f"❌ Ошибка: {str(e)}")
    # После сохранения возвращаем в главное меню
    await start(message, state)

@dp.callback_query(F.data == "generate_excel")
async def generate_excel(callback: types.CallbackQuery, state: FSMContext):
    # Личный отчёт: данные текущего пользователя
    user_id = callback.from_user.id
    cursor.execute("SELECT username, category, name, quantity FROM inventory WHERE user_id = ?", (user_id,))
    items = cursor.fetchall()
    if not items:
        await callback.answer("Нет данных!", show_alert=True)
        return
    wb = Workbook()
    ws = wb.active
    now = datetime.now().strftime("%Y-%m-%d_%H-%M")
    ws.title = f"Инвентаризация_{now}"
    ws.append(["Пользователь", "Категория", "Наименование", "Количество"])
    current_username = None
    current_category = None
    start_user_row = 2
    start_category_row = 2
    for row, (username, category, name, qty) in enumerate(items, start=2):
        if username != current_username:
            if current_username is not None:
                ws.merge_cells(f"A{start_user_row}:A{row-1}")
            current_username = username
            start_user_row = row
        if category != current_category:
            if current_category is not None:
                ws.merge_cells(f"B{start_category_row}:B{row-1}")
            current_category = category
            start_category_row = row
        ws.append([username if row == start_user_row else "",
                   category if row == start_category_row else "",
                   name, qty])
    if current_username:
        ws.merge_cells(f"A{start_user_row}:A{len(items)+1}")
    if current_category:
        ws.merge_cells(f"B{start_category_row}:B{len(items)+1}")
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    filename = f"inventory_{user_id}_{now}.xlsx"
    wb.save(filename)
    # Отправляем файл и удаляем его после отправки
    await bot.send_document(
        callback.from_user.id,
        document=types.BufferedInputFile(open(filename, "rb").read(), filename=filename),
        caption="✅ Ваш отчет готов!"
    )
    os.remove(filename)  # Удаление временного файла
    await callback.message.delete()
    await start(callback.message, state)
    await callback.answer()

@dp.callback_query(F.data == "generate_excel_all")
async def generate_excel_all(callback: types.CallbackQuery, state: FSMContext):
    # Общий отчёт: данные всех пользователей
    cursor.execute("SELECT username, category, name, quantity FROM inventory")
    items = cursor.fetchall()
    if not items:
        await callback.answer("Нет данных!", show_alert=True)
        return
    wb = Workbook()
    ws = wb.active
    now = datetime.now().strftime("%Y-%m-%d_%H-%M")
    ws.title = f"Общая_Инвентаризация_{now}"
    ws.append(["Пользователь", "Категория", "Наименование", "Количество"])
    for row in items:
        ws.append(list(row))
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    filename = f"общая_{now}.xlsx"
    wb.save(filename)
    await bot.send_document(
        callback.from_user.id,
        document=types.BufferedInputFile(open(filename, "rb").read(), filename=filename),
        caption="✅ Отчет общей инвентаризации готов!"
    )
    os.remove(filename)  # Удаление временного файла
    await callback.message.delete()
    await start(callback.message, state)
    await callback.answer()

# ========= Редактирование записей =========

@dp.callback_query(F.data == "edit_menu")
async def show_edit_menu(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    cursor.execute("SELECT rowid, category, name, quantity FROM inventory WHERE user_id = ?", (user_id,))
    records = cursor.fetchall()
    if not records:
        return await callback.answer("Нет записей для редактирования!", show_alert=True)
    builder = InlineKeyboardBuilder()
    for rowid, category, name, quantity in records:
        button_text = f"{category} - {name} ({quantity})"
        builder.button(text=button_text, callback_data=f"edit:{rowid}")
    builder.button(text="🔙 Назад", callback_data="back_to_main")
    builder.adjust(1)
    await callback.message.edit_text("Выберите запись для редактирования:", reply_markup=builder.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("edit:"))
async def edit_record(callback: types.CallbackQuery, state: FSMContext):
    rowid = callback.data.split(":", 1)[1]
    await state.update_data(edit_rowid=rowid)
    await callback.message.edit_text("Введите новые данные для записи в формате:\n<Новое название> <Новое количество>\nНапример: Виски_Jack 4.0")
    await state.set_state(InventoryState.waiting_for_edit_input)
    await callback.answer()

@dp.message(InventoryState.waiting_for_edit_input)
async def process_edit_input(message: types.Message, state: FSMContext):
    data = await state.get_data()
    rowid = data.get("edit_rowid")
    if not rowid:
        await message.answer("❌ Не удалось определить запись для редактирования.")
        return
    parts = message.text.strip().rsplit(" ", 1)
    if len(parts) != 2:
        return await message.answer("❌ Неверный формат. Введите данные в формате: <Новое название> <Новое количество>")
    new_name, quantity_str = parts[0].strip(), parts[1].replace(',', '.')
    try:
        new_quantity = float(quantity_str)
    except ValueError:
        return await message.answer("❌ Неверное значение количества. Попробуйте снова.")
    try:
        cursor.execute("UPDATE inventory SET name = ?, quantity = ? WHERE rowid = ?", (new_name, new_quantity, rowid))
        conn.commit()
        await message.answer("✅ Запись успешно обновлена!")
    except Exception as e:
        logging.error(f"Ошибка обновления записи: {e}")
        await message.answer(f"❌ Ошибка обновления записи: {e}")
    await state.clear()
    await start(message, state)

# ========= Удаление записей =========

@dp.callback_query(F.data == "delete_menu")
async def show_delete_menu(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    cursor.execute("SELECT rowid, category, name, quantity FROM inventory WHERE user_id = ?", (user_id,))
    records = cursor.fetchall()
    if not records:
        return await callback.answer("Нет записей для удаления!", show_alert=True)
    builder = InlineKeyboardBuilder()
    for rowid, category, name, quantity in records:
        button_text = f"{category} - {name} ({quantity})"
        builder.button(text=button_text, callback_data=f"delete:{rowid}")
    builder.button(text="🔙 Назад", callback_data="back_to_main")
    builder.adjust(1)
    await callback.message.edit_text("Выберите запись для удаления:", reply_markup=builder.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("delete:"))
async def confirm_delete(callback: types.CallbackQuery):
    rowid = callback.data.split(":", 1)[1]
    builder = InlineKeyboardBuilder()
    builder.button(text="Да", callback_data=f"confirm_delete:{rowid}")
    builder.button(text="Нет", callback_data="delete_cancel")
    builder.adjust(2)
    await callback.message.edit_text("Вы действительно хотите удалить выбранную запись?", reply_markup=builder.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("confirm_delete:"))
async def delete_record(callback: types.CallbackQuery):
    rowid = callback.data.split(":", 1)[1]
    try:
        cursor.execute("DELETE FROM inventory WHERE rowid = ?", (rowid,))
        conn.commit()
        await callback.message.edit_text("✅ Запись удалена.")
    except Exception as e:
        logging.error(f"Ошибка удаления записи: {e}")
        await callback.message.edit_text(f"❌ Ошибка удаления записи: {e}")
    await callback.answer()

@dp.callback_query(F.data == "delete_cancel")
async def delete_cancel(callback: types.CallbackQuery):
    await callback.message.edit_text("❌ Удаление отменено.")
    await callback.answer()

async def on_startup(bot: Bot) -> None:
    await bot.set_webhook(WEBHOOK_URL)

async def on_shutdown(bot: Bot) -> None:
    await bot.delete_webhook()

def main():
    app = web.Application()
    webhook_requests_handler = SimpleRequestHandler(
        dispatcher=dp,
        bot=bot,
    )
    webhook_requests_handler.register(app, path="/webhook")
    setup_application(app, dp, bot=bot)
    web.run_app(app, host="0.0.0.0", port=PORT)

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    WEBHOOK_URL = os.getenv("WEBHOOK_URL")
    PORT = int(os.getenv("PORT", 10000))
    main()
